import json
import re
import argparse
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass

import yaml
import unicodedata
import pandas as pd
from openpyxl import Workbook
from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# CONFIGURATION LOADER
# =============================================================================

def load_config(config_path: str) -> dict:
    """Load configuration from YAML file."""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


# =============================================================================
# DATA LOADERS
# =============================================================================

def load_papers_from_folder(folder_path: str) -> Dict[str, dict]:
    """Load all papers from JSON files in a folder."""
    papers = {}
    folder = Path(folder_path)
    
    for json_file in sorted(folder.glob('*.json')):
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if 'papers' in data:
                for paper in data['papers']:
                    papers[str(paper['paper_id'])] = paper
    
    return papers


# =============================================================================
# NORMALIZATION FUNCTIONS
# =============================================================================

def normalize_doi(doi: Optional[str], config: dict) -> Optional[str]:
    """Normalize DOI by removing URL prefixes."""
    if doi is None:
        return None
    
    doi = str(doi).strip()
    if not doi:
        return None
    
    if config['scoring_rules']['doi']['normalize']:
        prefixes = ['https://doi.org/', 'http://doi.org/', 
                    'http://dx.doi.org/', 'https://dx.doi.org/']
        for prefix in prefixes:
            doi = doi.replace(prefix, '')
    
    if not config['scoring_rules']['doi']['case_sensitive']:
        doi = doi.lower()
    
    return doi if doi else None


def normalize_country(country: str, config: dict) -> str:
    """Normalize country name using config mapping."""
    country = country.lower().strip()
    norm_map = config.get('country_normalization', {})
    return norm_map.get(country, country)


def normalize_keyword(keyword: Optional[str], config: dict) -> Optional[str]:
    """Normalize keyword for comparison."""
    if not keyword:
        return None
    
    kw = str(keyword).strip()
    
    # Normalize quotes
    # kw = re.sub(r'[''`’?]', "", kw)
    kw = re.sub(r'[''`’?]', "'", kw)
    kw = re.sub(r'[""]', '"', kw)
    
    if not config['scoring_rules']['keywords']['case_sensitive']:
        kw = kw.lower()
    
    return kw if kw else None


def normalize_title(title: Optional[str], config: dict) -> Optional[str]:
    """Normalize title for comparison."""
    if title is None:
        return None
    
    t = str(title).strip()
    
    if config['scoring_rules']['title']['normalize_encoding']:
        # Normalize whitespace
        t = re.sub(r'\s+', ' ', t)
        # Remove special quote variations
        t = re.sub(r'[''`""’?]', '', t)
        # Remove punctuation for comparison
        t = re.sub(r'[^\w\s]', '', t)
    
    if not config['scoring_rules']['title']['case_sensitive']:
        t = t.lower()
    
    return t if t else None


def normalize_author_name(name: str, config: dict) -> str:
    """Normalize full author name for comparison."""
    name = str(name)
    
    if config['scoring_rules']['authors']['ignore_titles']:
        # Remove common titles and degrees
        titles = r'(,?\s*(MD|PhD|MRCP|MBBS|FAMS|FRCPEd|MRCPGP|DCH|DRCOG|ChB|MS|PhDc|MDa|MDb|MDe|MDf|Dr|Prof|Jr|Sr|II|III)\.?)'
        name = re.sub(titles, '', name, flags=re.I)
    
    # Remove numbers and special markers
    name = re.sub(r'[\d*†,]', '', name)
    
    # Unicode normalization: é -> e
    name = unicodedata.normalize('NFKD', name)
    name = ''.join(c for c in name if not unicodedata.combining(c))
    
    # Normalize whitespace and hyphens
    name = name.replace('-', ' ')
    name = re.sub(r'\s+', ' ', name)
    
    # Lowercase and strip
    name = name.lower().strip()
    
    return name


# =============================================================================
# SCORING FUNCTIONS
# =============================================================================

@dataclass
class FieldResult:
    """Result of evaluating a single field."""
    score: int
    explanation: str = ""
    # track TP, FP, FN for calculation
    tp: int = 0  # True Positive: GT exists and correctly extracted
    fp: int = 0  # False Positive: GT is null but model extracted something
    fn: int = 0  # False Negative: GT exists but model failed to extract


def score_doi(gt_val: Optional[str], exp_val: Optional[str], config: dict) -> FieldResult:
    gt_doi = normalize_doi(gt_val, config)
    exp_doi = normalize_doi(exp_val, config)
    
    if gt_doi is None and exp_doi is None:
        # True Negative - no GT, no extraction (not counted in P/R)
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    elif gt_doi is None and exp_doi is not None:
        # False Positive - hallucinated
        return FieldResult(score=0, explanation=f"GT DOI is null but detected '{exp_val}'", tp=0, fp=1, fn=0)
    elif gt_doi is not None and exp_doi is None:
        # False Negative - missed
        return FieldResult(score=0, explanation="GT DOI not detected", tp=0, fp=0, fn=1)
    elif gt_doi == exp_doi:
        # True Positive - correct extraction
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        # Wrong extraction - both missed correct (FN) AND hallucinated wrong (FP)
        return FieldResult(score=0, explanation="DOI mismatch", tp=0, fp=1, fn=1)


def score_title(gt_val: Optional[str], exp_val: Optional[str], config: dict) -> FieldResult:
    if gt_val is None and exp_val is None:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    elif gt_val is None:
        return FieldResult(score=0, explanation="GT Title is null but detected", tp=0, fp=1, fn=0)
    elif exp_val is None:
        return FieldResult(score=0, explanation="Title not detected", tp=0, fp=0, fn=1)
    
    gt_norm = normalize_title(gt_val, config)
    exp_norm = normalize_title(exp_val, config)
    
    if gt_norm == exp_norm or gt_norm in exp_norm or exp_norm in gt_norm:
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        return FieldResult(score=0, explanation="Title mismatch", tp=0, fp=1, fn=1)


def score_year(gt_val, exp_val, config: dict) -> FieldResult:
    if gt_val is None and exp_val is None:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    elif gt_val is None:
        return FieldResult(score=0, explanation=f"GT Year is null but detected {exp_val}", tp=0, fp=1, fn=0)
    elif exp_val is None:
        return FieldResult(score=0, explanation=f"Year not detected (GT={gt_val})", tp=0, fp=0, fn=1)
    
    try:
        if int(gt_val) == int(exp_val):
            return FieldResult(score=1, tp=1, fp=0, fn=0)
        else:
            return FieldResult(score=0, explanation=f"Year mismatch: GT={gt_val}, Exp={exp_val}", tp=0, fp=1, fn=1)
    except (ValueError, TypeError):
        if str(gt_val) == str(exp_val):
            return FieldResult(score=1, tp=1, fp=0, fn=0)
        else:
            return FieldResult(score=0, explanation=f"Year mismatch: GT={gt_val}, Exp={exp_val}", tp=0, fp=1, fn=1)


def score_authors(gt_list: List[str], exp_list: List[str], config: dict) -> FieldResult:
    gt_list = gt_list or []
    exp_list = exp_list or []
    
    if not gt_list and not exp_list:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    elif not gt_list:
        return FieldResult(score=0, explanation=f"GT Authors empty but detected {len(exp_list)}", tp=0, fp=1, fn=0)
    elif not exp_list:
        return FieldResult(score=0, explanation="Authors not detected", tp=0, fp=0, fn=1)
    
    # Normalize all names
    gt_names = [normalize_author_name(a, config) for a in gt_list]
    exp_names = [normalize_author_name(a, config) for a in exp_list]
    
    # Filter empty names
    gt_names = [n for n in gt_names if n]
    exp_names = [n for n in exp_names if n]
    
    if not gt_names:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    
    # === STEP 1: Strict count check ===
    if len(gt_names) != len(exp_names):
        return FieldResult(
            score=0,
            explanation=f"Authors: count mismatch (GT={len(gt_names)}, Exp={len(exp_names)})",
            tp=0, fp=1, fn=1
        )
    
    # === STEP 2: Fuzzy matching (only for unicode tolerance) ===
    fuzzy_threshold = config['scoring_rules']['authors'].get('fuzzy_threshold', 0.90)
    
    matched_count = 0
    unmatched = []
    exp_used = [False] * len(exp_names)  # Track which Exp names are already matched
    
    for gt_name in gt_names:
        found = False
        best_match_idx = -1
        best_match_score = 0
        
        for idx, exp_name in enumerate(exp_names):
            if exp_used[idx]:
                continue
            
            # Exact match
            if gt_name == exp_name:
                found = True
                best_match_idx = idx
                break
            
            # Fuzzy match
            similarity = SequenceMatcher(None, gt_name, exp_name).ratio()
            if similarity >= fuzzy_threshold and similarity > best_match_score:
                best_match_score = similarity
                best_match_idx = idx
                found = True
        
        if found and best_match_idx >= 0:
            matched_count += 1
            exp_used[best_match_idx] = True  # Mark as used
        else:
            unmatched.append(gt_name)
    
    # === STEP 3: Threshold check ===
    overlap = matched_count / len(gt_names)
    threshold = config['scoring_rules']['authors'].get('threshold', 0.7)
    
    if overlap >= threshold:
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        max_display = config['output']['descriptions']['max_missing_authors']
        missing_display = unmatched[:max_display]
        if len(unmatched) > max_display:
            explanation = f"Authors: {overlap*100:.0f}% match ({matched_count}/{len(gt_names)}), unmatched: {missing_display}..."
        else:
            explanation = f"Authors: {overlap*100:.0f}% match ({matched_count}/{len(gt_names)}), unmatched: {missing_display}"
        return FieldResult(score=0, explanation=explanation, tp=0, fp=1, fn=1)


def score_countries(gt_list: List[str], exp_list: List[str], config: dict) -> FieldResult:
    gt_list = gt_list or []
    exp_list = exp_list or []
    
    gt_norm = set([normalize_country(c, config) for c in gt_list if c])
    exp_norm = set([normalize_country(c, config) for c in exp_list if c])
    
    if not gt_norm:
        if exp_norm:
            return FieldResult(score=1, tp=0, fp=0, fn=0)  
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    
    matched = gt_norm & exp_norm
    match_ratio = len(matched) / len(gt_norm)
    
    threshold = config['scoring_rules']['countries'].get('threshold', 0.75)
    
    if match_ratio >= threshold:
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        missing = gt_norm - exp_norm
        max_display = config['output']['descriptions']['max_missing_countries']
        missing_display = list(missing)[:max_display]
        if len(missing) > max_display:
            explanation = f"Countries: {match_ratio*100:.0f}% match ({len(matched)}/{len(gt_norm)}), missing: {', '.join(missing_display)}..."
        else:
            explanation = f"Countries: {match_ratio*100:.0f}% match ({len(matched)}/{len(gt_norm)}), missing: {missing_display}"
        return FieldResult(score=0, explanation=explanation, tp=0, fp=1, fn=1)

def score_purpose(gt_val: Optional[str], exp_val: Optional[str], config: dict) -> FieldResult:
    if gt_val is None and exp_val is None:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    elif gt_val is None:
        return FieldResult(score=0, explanation="GT Purpose is null but detected", tp=0, fp=1, fn=0)
    elif exp_val is None:
        return FieldResult(score=0, explanation="Purpose not detected", tp=0, fp=0, fn=1)
    
    stopwords = set(config.get('stopwords', []))
    
    gt_words = set(str(gt_val).lower().split()) - stopwords
    exp_words = set(str(exp_val).lower().split()) - stopwords
    
    if not gt_words:
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    
    overlap = len(gt_words & exp_words) / len(gt_words)
    threshold = config['scoring_rules']['purpose_of_work']['threshold']
    
    if overlap >= threshold:
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        return FieldResult(score=0, explanation=f"Purpose: {overlap*100:.0f}% semantic overlap", tp=0, fp=1, fn=1)


def score_keywords(gt_list: List[str], exp_list: List[str], config: dict) -> FieldResult:
    gt_list = gt_list or []
    exp_list = exp_list or []
    
    gt_norm = set([normalize_keyword(k, config) for k in gt_list if k])
    exp_norm = set([normalize_keyword(k, config) for k in exp_list if k])
    
    exp_expanded = set()
    for k in exp_norm:
        if k and ';' in k:
            exp_expanded.update([x.strip() for x in k.split(';') if x.strip()])
        elif k:
            exp_expanded.add(k)
    exp_norm = exp_expanded
    
    gt_norm = {k for k in gt_norm if k}
    exp_norm = {k for k in exp_norm if k}
    
    if not gt_norm:
        if exp_norm and config['scoring_rules']['keywords']['handle_empty_gt'] == 'score_zero_if_detected':
            return FieldResult(score=0, explanation=f"Keywords: GT empty but detected {len(exp_norm)}", tp=0, fp=1, fn=0)
        return FieldResult(score=1, tp=0, fp=0, fn=0)
    
    matched = gt_norm & exp_norm
    match_ratio = len(matched) / len(gt_norm)
    
    threshold = config['scoring_rules']['keywords'].get('threshold', 0.5)
    
    if match_ratio >= threshold:
        return FieldResult(score=1, tp=1, fp=0, fn=0)
    else:
        missing = gt_norm - exp_norm
        max_display = config['output']['descriptions']['max_missing_keywords']
        missing_display = list(missing)[:max_display]
        if len(missing) > max_display:
            explanation = f"Keywords: {match_ratio*100:.0f}% match ({len(matched)}/{len(gt_norm)}), missing: {', '.join(missing_display)}..."
        else:
            explanation = f"Keywords: {match_ratio*100:.0f}% match ({len(matched)}/{len(gt_norm)}), missing: {missing_display}"
        return FieldResult(score=0, explanation=explanation, tp=0, fp=1, fn=1)

# =============================================================================
# MAIN EVALUATION FUNCTION
# =============================================================================

def evaluate_paper(gt: dict, exp: dict, config: dict) -> dict:
    """Evaluate a single paper and return scores with explanations."""
    results = {}
    explanations = []
    
    # Track TP, FP, FN for this paper (overall)
    total_tp = 0
    total_fp = 0
    total_fn = 0
    
    # DOI
    doi_result = score_doi(gt.get('doi'), exp.get('doi'), config)
    results['DOI'] = doi_result.score
    results['_DOI_tp'] = doi_result.tp
    results['_DOI_fp'] = doi_result.fp
    results['_DOI_fn'] = doi_result.fn
    total_tp += doi_result.tp
    total_fp += doi_result.fp
    total_fn += doi_result.fn
    if doi_result.explanation:
        explanations.append(doi_result.explanation)
    
    # Title
    title_result = score_title(gt.get('title'), exp.get('title'), config)
    results['Title'] = title_result.score
    results['_Title_tp'] = title_result.tp
    results['_Title_fp'] = title_result.fp
    results['_Title_fn'] = title_result.fn
    total_tp += title_result.tp
    total_fp += title_result.fp
    total_fn += title_result.fn
    if title_result.explanation:
        explanations.append(title_result.explanation)
    
    # Year
    year_result = score_year(gt.get('published_year'), exp.get('published_year'), config)
    results['Published year'] = year_result.score
    results['_Published year_tp'] = year_result.tp
    results['_Published year_fp'] = year_result.fp
    results['_Published year_fn'] = year_result.fn
    total_tp += year_result.tp
    total_fp += year_result.fp
    total_fn += year_result.fn
    if year_result.explanation:
        explanations.append(year_result.explanation)
    
    # Authors
    authors_result = score_authors(
        gt.get('author_list', []), 
        exp.get('author_list', []), 
        config
    )
    results['Authors'] = authors_result.score
    results['_Authors_tp'] = authors_result.tp
    results['_Authors_fp'] = authors_result.fp
    results['_Authors_fn'] = authors_result.fn
    total_tp += authors_result.tp
    total_fp += authors_result.fp
    total_fn += authors_result.fn
    if authors_result.explanation:
        explanations.append(authors_result.explanation)
    
    # Countries
    countries_result = score_countries(
        gt.get('countries', []), 
        exp.get('countries', []), 
        config
    )
    results['Countries'] = countries_result.score
    results['_Countries_tp'] = countries_result.tp
    results['_Countries_fp'] = countries_result.fp
    results['_Countries_fn'] = countries_result.fn
    total_tp += countries_result.tp
    total_fp += countries_result.fp
    total_fn += countries_result.fn
    if countries_result.explanation:
        explanations.append(countries_result.explanation)
    
    # Purpose
    purpose_result = score_purpose(
        gt.get('purpose_of_work'), 
        exp.get('purpose_of_work'), 
        config
    )
    results['Purpose of work'] = purpose_result.score
    results['_Purpose of work_tp'] = purpose_result.tp
    results['_Purpose of work_fp'] = purpose_result.fp
    results['_Purpose of work_fn'] = purpose_result.fn
    total_tp += purpose_result.tp
    total_fp += purpose_result.fp
    total_fn += purpose_result.fn
    if purpose_result.explanation:
        explanations.append(purpose_result.explanation)
    
    # Keywords
    keywords_result = score_keywords(
        gt.get('keywords', []), 
        exp.get('keywords', []), 
        config
    )
    results['keywords'] = keywords_result.score
    results['_keywords_tp'] = keywords_result.tp
    results['_keywords_fp'] = keywords_result.fp
    results['_keywords_fn'] = keywords_result.fn
    total_tp += keywords_result.tp
    total_fp += keywords_result.fp
    total_fn += keywords_result.fn
    if keywords_result.explanation:
        explanations.append(keywords_result.explanation)
    
    # Calculate totals
    fields = ['DOI', 'Title', 'Published year', 'Authors', 'Countries', 'Purpose of work', 'keywords']
    results['Correct /7'] = sum(results[f] for f in fields)
    results['Wrong'] = 7 - results['Correct /7']
    results['Accuracy(%)'] = results['Correct /7'] / 7
    
    # Calculate per-paper Precision, Recall, F1
    precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 1.0
    recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 1.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    
    results['Precision(%)'] = precision
    results['Recall(%)'] = recall
    results['F1(%)'] = f1
    
    # Store TP, FP, FN for aggregation
    results['_tp'] = total_tp
    results['_fp'] = total_fp
    results['_fn'] = total_fn
    
    results['description of failed extractions'] = '; '.join(explanations) if explanations else ''
    
    return results


def evaluate_experiment(gt_papers: Dict[str, dict], exp_papers: Dict[str, dict], config: dict) -> pd.DataFrame:
    """Evaluate all papers in an experiment."""
    results = []
    
    fields = ['DOI', 'Title', 'Published year', 'Authors', 'Countries', 'Purpose of work', 'keywords']
    
    for paper_id in sorted(gt_papers.keys(), key=lambda x: int(x)):
        gt = gt_papers[paper_id]
        exp = exp_papers.get(paper_id)
        
        # Handle missing papers - score all fields as 0
        if exp is None:
            print(f"WARNING: Paper {paper_id} missing from experiment results - scoring all fields as 0")
            row = {
                'ID': int(paper_id),
                'DOI': 0, 'Title': 0, 'Published year': 0, 'Authors': 0,
                'Countries': 0, 'Purpose of work': 0, 'keywords': 0,
                'Correct /7': 0, 'Wrong': 7, 'Accuracy(%)': 0.0,
                'Precision(%)': 0.0, 'Recall(%)': 0.0, 'F1(%)': 0.0,
                'description of failed extractions': 'Paper not found in experiment results',
                '_tp': 0, '_fp': 0, '_fn': 7
            }
            for field in fields:
                row[f'_{field}_tp'] = 0
                row[f'_{field}_fp'] = 0
                row[f'_{field}_fn'] = 1
            results.append(row)
            continue
        
        row = {'ID': int(paper_id)}
        row.update(evaluate_paper(gt, exp, config))
        results.append(row)
    
    # Create DataFrame with ordered columns
    display_columns = ['ID', 'DOI', 'Title', 'Published year', 'Authors', 'Countries', 
                       'Purpose of work', 'keywords', 'Correct /7', 'Wrong', 
                       'Accuracy(%)', 'Precision(%)', 'Recall(%)', 'F1(%)', 
                       'description of failed extractions']
    
    # Internal columns for metrics calculation
    fields = ['DOI', 'Title', 'Published year', 'Authors', 'Countries', 'Purpose of work', 'keywords']
    internal_columns = ['_tp', '_fp', '_fn']
    for field in fields:
        internal_columns.extend([f'_{field}_tp', f'_{field}_fp', f'_{field}_fn'])
    
    df = pd.DataFrame(results)
    
    # Store TP/FP/FN totals as DataFrame attributes for later use
    df.attrs['total_tp'] = df['_tp'].sum()
    df.attrs['total_fp'] = df['_fp'].sum()
    df.attrs['total_fn'] = df['_fn'].sum()
    
    return df[display_columns + internal_columns]


def calculate_metrics(df: pd.DataFrame) -> dict:
    """Calculate Precision, Recall, F1 from aggregated TP, FP, FN."""
    total_tp = df['_tp'].sum()
    total_fp = df['_fp'].sum()
    total_fn = df['_fn'].sum()
    
    # Precision = TP / (TP + FP)
    precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0.0
    
    # Recall = TP / (TP + FN)
    recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0.0
    
    # F1 = 2 * (Precision * Recall) / (Precision + Recall)
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0
    
    # Accuracy (existing)
    accuracy = df['Correct /7'].sum() / (len(df) * 7)
    
    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'total_tp': int(total_tp),
        'total_fp': int(total_fp),
        'total_fn': int(total_fn)
    }


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def create_excel_output(df: pd.DataFrame, output_path: str, config: dict, experiment_name: str = ""):
    """Create formatted Excel output."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"
    
    # Get colors from config
    output_config = config['output']['excel']
    header_color = output_config['header_color']
    score_1_color = output_config['score_1_color']
    score_0_color = output_config['score_0_color']
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    score_0_fill = PatternFill(start_color=score_0_color, end_color=score_0_color, fill_type='solid')
    score_1_fill = PatternFill(start_color=score_1_color, end_color=score_1_color, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Filter out internal columns for display
    display_columns = [c for c in df.columns if not c.startswith('_')]
    df_display = df[display_columns]
    
    # Write headers
    headers = list(df_display.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data
    score_columns = [2, 3, 4, 5, 6, 7, 8]  # DOI through keywords
    
    for row_idx, row in enumerate(df_display.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Color code binary scores
            if output_config['color_code_scores'] and col_idx in score_columns:
                if value == 0:
                    cell.fill = score_0_fill
                elif value == 1:
                    cell.fill = score_1_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Center certain columns
            if col_idx in [1, 9, 10]:
                cell.alignment = Alignment(horizontal='center')
            
            # Format percentage columns (Accuracy, Precision, Recall, F1)
            if col_idx in [11, 12, 13, 14]:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = {'A': 6, 'B': 6, 'C': 6, 'D': 12, 'E': 8, 'F': 10, 
                     'G': 14, 'H': 10, 'I': 10, 'J': 8, 'K': 12, 'L': 12, 
                     'M': 10, 'N': 8, 'O': 80}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add summary sheet 
    if output_config['include_summary_sheet']:
        add_summary_sheet(wb, df, experiment_name)
    
    wb.save(output_path)


def add_summary_sheet(wb: Workbook, df: pd.DataFrame, experiment_name: str):
    ws = wb.create_sheet("Summary")
    
    fields = ['DOI', 'Title', 'Published year', 'Authors', 'Countries', 'Purpose of work', 'keywords']
    
    # Calculate overall metrics
    metrics = calculate_metrics(df)
    
    summary_data = [
        ["EVALUATION SUMMARY", "", "", "", "", "", ""],
        ["Experiment", experiment_name, "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Total Papers", len(df), "", "", "", "", ""],
        ["Total Fields Evaluated", len(df) * 7, "", "", "", "", ""],
        ["Total Correct", int(df['Correct /7'].sum()), "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["OVERALL METRICS", "", "", "", "", "", ""],
        ["Metric", "Value", "", "", "", "", ""],
        ["Accuracy", f"{metrics['accuracy']*100:.2f}%", "", "", "", "", ""],
        ["Precision", f"{metrics['precision']*100:.2f}%", "", "", "", "", ""],
        ["Recall", f"{metrics['recall']*100:.2f}%", "", "", "", "", ""],
        ["F1 Score", f"{metrics['f1']*100:.2f}%", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["PER-FIELD PERFORMANCE", "", "", "", "", "", ""],
        ["Field", "Correct", "Total", "Accuracy", "Precision", "Recall", "F1"],
    ]
    
    # Calculate per-field metrics using stored TP/FP/FN
    for field in fields:
        field_tp = int(df[f'_{field}_tp'].sum())
        field_fp = int(df[f'_{field}_fp'].sum())
        field_fn = int(df[f'_{field}_fn'].sum())
        
        correct = int(df[field].sum())
        total = len(df)
        
        # Calculate metrics
        precision = field_tp / (field_tp + field_fp) if (field_tp + field_fp) > 0 else 1.0
        recall = field_tp / (field_tp + field_fn) if (field_tp + field_fn) > 0 else 1.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
        accuracy = correct / total
        
        summary_data.append([
            field,
            correct,
            total,
            f"{accuracy*100:.1f}%",
            f"{precision*100:.1f}%",
            f"{recall*100:.1f}%",
            f"{f1*100:.1f}%"
        ])
    
    summary_data.extend([
        ["", "", "", "", "", "", ""],
        ["SCORE DISTRIBUTION", "", "", "", "", "", ""],
        ["Score", "Count", "", "", "", "", ""],
        ["7/7 (Perfect)", len(df[df['Correct /7'] == 7]), "", "", "", "", ""],
        ["6/7", len(df[df['Correct /7'] == 6]), "", "", "", "", ""],
        ["5/7", len(df[df['Correct /7'] == 5]), "", "", "", "", ""],
        ["<5/7", len(df[df['Correct /7'] < 5]), "", "", "", "", ""],
    ])
    
    # Style settings
    bold_rows = [1, 8, 15, 24]  
    header_rows = [9, 16, 25]  
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in bold_rows:
                cell.font = Font(bold=True, size=12)
            if row_idx in header_rows:
                cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Evaluate LLM metadata extraction against ground truth',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python evaluate.py --gt ./gt --exp ./llama_results --output ./results/llama.xlsx
  python evaluate.py --config custom_config.yaml --gt ./gt --exp ./exp --output result.xlsx
        """
    )
    
    parser.add_argument('--config', default='config.yaml', 
                        help='Path to configuration file (default: config.yaml)')
    parser.add_argument('--gt', required=True,
                        help='Path to ground truth folder containing JSON files')
    parser.add_argument('--exp', required=True,
                        help='Path to experiment results folder')
    parser.add_argument('--output', required=True,
                        help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Load configuration
    config = load_config(args.config)
    print(f"Loaded configuration from: {args.config}")
    
    # Create output directory if needed
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Load papers
    gt_papers = load_papers_from_folder(args.gt)
    exp_papers = load_papers_from_folder(args.exp)
    
    print(f"Loaded {len(gt_papers)} GT papers, {len(exp_papers)} experiment papers")
    
    # Evaluate
    df = evaluate_experiment(gt_papers, exp_papers, config)
    
    exp_name = Path(args.exp).name
    create_excel_output(df, str(output_path), config, exp_name)
    
    # Calculate and print metrics
    metrics = calculate_metrics(df)
    
    print(f"\n{'='*60}")
    print("EVALUATION COMPLETE")
    print(f"{'='*60}")
    print(f"Output saved to: {output_path}")
    print(f"\nOVERALL METRICS:")
    print(f"  Accuracy:  {metrics['accuracy']*100:.2f}%")
    print(f"  Precision: {metrics['precision']*100:.2f}%")
    print(f"  Recall:    {metrics['recall']*100:.2f}%")
    print(f"  F1 Score:  {metrics['f1']*100:.2f}%")
    print(f"\nPer-field accuracy:")
    for field in ['DOI', 'Title', 'Published year', 'Authors', 'Countries', 'Purpose of work', 'keywords']:
        print(f"  {field}: {int(df[field].sum())}/{len(df)} ({df[field].sum()/len(df)*100:.1f}%)")


if __name__ == '__main__':
    main()